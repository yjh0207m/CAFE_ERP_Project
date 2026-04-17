"""
excel.py — 재무분석 엑셀 보고서 생성 (openpyxl)
시트 구성:
  1. 손익계산서
  2. 재무상태표
  3. 자본변동표
  4. 분개장
  5. 수익예측
  6. 재고예측
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from config import EXCEL_OUTPUT_DIR

# ── 공통 스타일 ───────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="3B5998")
SUBHEADER_FILL = PatternFill("solid", fgColor="8EA9D8")
SUBTOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
TITLE_FILL     = PatternFill("solid", fgColor="2E4053")
ALT_FILL       = PatternFill("solid", fgColor="EBF5FB")

HEADER_FONT    = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TITLE_FONT     = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
BODY_FONT      = Font(name="맑은 고딕", size=10)
SUBTOTAL_FONT  = Font(name="맑은 고딕", bold=True, size=10)
NEGATIVE_FONT  = Font(name="맑은 고딕", size=10, color="C0392B")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

NUM_FMT = '#,##0'


def _style(ws, row, col, value, fill=None, font=None,
           align=None, border=True, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = THIN_BORDER
    if num_fmt: cell.number_format = num_fmt
    return cell


def _title_row(ws, title: str, cols: int, row: int = 1):
    """시트 최상단 타이틀 병합 행"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill      = TITLE_FILL
    cell.font      = TITLE_FONT
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def _header_row(ws, headers: list, row: int):
    for c, h in enumerate(headers, 1):
        _style(ws, row, c, h, fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)


def _money(ws, row, col, value, is_subtotal=False):
    fill = SUBTOTAL_FILL if is_subtotal else None
    font = SUBTOTAL_FONT if is_subtotal else BODY_FONT
    if isinstance(value, (int, float)) and value < 0:
        font = NEGATIVE_FONT
    cell = _style(ws, row, col, value, fill=fill, font=font, align=RIGHT)
    cell.number_format = NUM_FMT


# ═══════════════════════════════════════════════════════════════════
# ① 손익계산서
# ═══════════════════════════════════════════════════════════════════
def _write_income_sheet(ws, income: dict):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20

    period = income.get("period", "")
    _title_row(ws, f"손익계산서  ({period})", 3)
    _header_row(ws, ["항  목", "금  액 (원)", "비  고"], 2)

    r = 3

    def item(label, amount, note="", subtotal=False, highlight=False):
        nonlocal r
        fill = SUBHEADER_FILL if highlight else (SUBTOTAL_FILL if subtotal else None)
        font = SUBTOTAL_FONT  if (highlight or subtotal) else BODY_FONT
        _style(ws, r, 1, label, fill=fill, font=font, align=LEFT)
        if isinstance(amount, (int, float)):
            # highlight일 때 B열도 동일한 fill 직접 적용
            cell = ws.cell(row=r, column=2, value=amount)
            cell.fill          = fill or PatternFill()
            cell.font          = font
            cell.alignment     = RIGHT
            cell.border        = THIN_BORDER
            cell.number_format = NUM_FMT
        else:
            _style(ws, r, 2, amount, fill=fill, font=font, align=RIGHT)
        _style(ws, r, 3, note, fill=fill, font=font, align=LEFT)
        r += 1

    payroll = income.get("payroll", {})
    purch   = income.get("purchases", {})
    exp     = income.get("expenses", {})
    rev     = income.get("revenue", {})
    summary = income.get("summary", {})

    # ── 급여 지출 ─────────────────────────────────────────────────
    item("  매니저 급여",  payroll.get("매니저", 0))
    item("  점장 급여",    payroll.get("점장",   0))
    item("  스탭 급여",    payroll.get("스탭",   0))
    item("  인센티브",     payroll.get("인센티브", 0))
    item("  급여 소계",    payroll.get("급여소계", 0), subtotal=True)

    # ── 발주 비용 ─────────────────────────────────────────────────
    item("  원두 발주",      purch.get("원두",    0))
    item("  유제품 발주",    purch.get("유제품",  0))
    item("  시럽/소스 발주", purch.get("시럽/소스", 0))
    item("  파우더 발주",    purch.get("파우더",  0))
    item("  차류 발주",      purch.get("차류",    0))
    item("  소모품 발주",    purch.get("소모품",  0))
    item("  기타 발주",      purch.get("기타",    0))
    item("  발주 소계",      purch.get("발주소계", 0), subtotal=True)

    # ── 기타 지출 ─────────────────────────────────────────────────
    item("  재료비",   exp.get("재료비",  0))
    item("  인건비",   exp.get("인건비",  0))
    item("  임대료",   exp.get("임대료",  0))
    item("  공과금",   exp.get("공과금",  0))
    item("  소모품",   exp.get("소모품",  0))
    item("  마케팅",   exp.get("마케팅",  0))
    item("  기타",     exp.get("기타",    0))
    item("  지출 소계", exp.get("지출소계", summary.get("지출소계", 0)), subtotal=True)
    item("  총 지출",  summary.get("총지출", 0), subtotal=True)

    # ── 매출 ──────────────────────────────────────────────────────
    item("  커피류 매출",         rev.get("커피류",          0))
    item("  논커피류 매출",       rev.get("논커피류",        0))
    item("  디저트 매출",         rev.get("디저트",          0))
    item("  스무디/프라푸치노 매출", rev.get("스무디/프라푸치노", 0))
    item("  티/한방 매출",        rev.get("티/한방",         0))
    item("  에이드 매출",         rev.get("에이드",          0))
    item("  베이커리 매출",       rev.get("베이커리",        0))
    item("  샌드위치/브런치 매출", rev.get("샌드위치/브런치", 0))
    item("  기타 매출",           rev.get("기타",            0))
    item("  총 매출", summary.get("총매출", 0), subtotal=True)

    # ── 당기순이익 ────────────────────────────────────────────────
    item("당기순이익", summary.get("당기순이익", 0), highlight=True)


# ═══════════════════════════════════════════════════════════════════
# ② 재무상태표
# ═══════════════════════════════════════════════════════════════════
def _write_balance_sheet(ws, balance: dict):
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    _title_row(ws, "재무상태표", 3)
    _header_row(ws, ["항  목", "당기말", "전기말"], 2)

    current  = balance.get("current", {})
    previous = balance.get("previous", {})

    rows = [
        ("현금",    False),
        ("재고",    False),
        ("자본합계", True),
        ("미착재고", False),
        ("부채합계", True),
        ("자산총액", True),
    ]

    for i, (label, is_sub) in enumerate(rows, 3):
        fill = SUBTOTAL_FILL if is_sub else (ALT_FILL if i % 2 == 0 else None)
        font = SUBTOTAL_FONT if is_sub else BODY_FONT
        _style(ws, i, 1, label, fill=fill, font=font, align=LEFT)
        curr_v = current.get(label, 0)
        prev_v = previous.get(label, 0)
        for col, val in [(2, curr_v), (3, prev_v)]:
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill or (ALT_FILL if i % 2 == 0 else PatternFill())
            cell.font   = font
            cell.alignment = RIGHT
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT


# ═══════════════════════════════════════════════════════════════════
# ③ 자본변동표
# ═══════════════════════════════════════════════════════════════════
def _write_capital_changes(ws, capital: dict):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    _title_row(ws, "자본변동표", 4)
    _header_row(ws, ["항  목", "현  금", "재  고", "자본합계"], 2)

    def row3(r, label, cash, inv, total, subtotal=False):
        fill = SUBTOTAL_FILL if subtotal else None
        font = SUBTOTAL_FONT if subtotal else BODY_FONT
        _style(ws, r, 1, label, fill=fill, font=font, align=LEFT)
        for c, v in [(2, cash), (3, inv), (4, total)]:
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill or PatternFill()
            cell.font = font
            cell.alignment = RIGHT
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT

    beg  = capital.get("기초자본", {})
    end  = capital.get("기말자본", {})
    cc   = capital.get("현금변동", 0)
    ic   = capital.get("재고변동", 0)
    tc   = capital.get("변동총액", 0)

    row3(3, "기초자본", beg.get("현금", 0), beg.get("재고", 0), beg.get("자본합계", 0))
    row3(4, "현금변동", cc, 0, cc)
    row3(5, "재고변동", 0, ic, ic)
    row3(6, "변동총액", cc, ic, tc, subtotal=True)
    row3(7, "기말자본", end.get("현금", 0), end.get("재고", 0), end.get("자본합계", 0), subtotal=True)


# ═══════════════════════════════════════════════════════════════════
# ④ 분개장
# ═══════════════════════════════════════════════════════════════════
def _write_journal(ws, journal: dict):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    _title_row(ws, f"분개장  ({journal.get('period', '')})", 6)
    _header_row(ws, ["날짜", "적요", "차변", "대변", "금액 (원)", "참조"], 2)

    entries = journal.get("entries", [])
    for i, e in enumerate(entries, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        for c, val in enumerate([
            e.get("date", ""), e.get("description", ""),
            e.get("debit", ""), e.get("credit", ""),
            e.get("amount", 0), e.get("ref", ""),
        ], 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill      = fill or PatternFill()
            cell.font      = BODY_FONT
            cell.alignment = RIGHT if c == 5 else CENTER if c in (1, 3, 4, 6) else LEFT
            cell.border    = THIN_BORDER
            if c == 5:
                cell.number_format = NUM_FMT


# ═══════════════════════════════════════════════════════════════════
# ⑤ 수익 예측
# ═══════════════════════════════════════════════════════════════════
def _write_revenue_forecast(ws, forecast: dict):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26

    history   = forecast.get("history",   [])
    forecasts = forecast.get("forecasts", [])
    summary   = forecast.get("summary",   {})

    model_used  = summary.get("model", "holt")
    model_label = "Prophet ML" if model_used == "prophet" else "Holt's 통계"
    title_text  = f"수익 예측  [예측 모델: {model_label}]"

    _title_row(ws, title_text, 5)
    _header_row(ws, ["기간", "예측 매출 (원)", "예측 지출 (원)", "예측 순이익 (원)", "매출 예측 범위 (95%)"], 2)

    r = 3
    # 실적 (회색 배경)
    for h in history:
        fill = PatternFill("solid", fgColor="D5D8DC")
        _style(ws, r, 1, h.get("label", ""),    fill=fill, font=BODY_FONT, align=CENTER)
        _money(ws, r, 2, h.get("revenue", 0));  ws.cell(row=r, column=2).fill = fill
        _money(ws, r, 3, h.get("expenses", 0)); ws.cell(row=r, column=3).fill = fill
        _money(ws, r, 4, h.get("net_income", 0)); ws.cell(row=r, column=4).fill = fill
        _style(ws, r, 5, "실적", fill=fill, font=BODY_FONT, align=CENTER)
        r += 1

    # 예측 (파란 배경)
    for fc in forecasts:
        fill = PatternFill("solid", fgColor="D6EAF8")
        _style(ws, r, 1, f"▶ {fc.get('label', '')} (예측)", fill=fill, font=SUBTOTAL_FONT, align=CENTER)
        _money(ws, r, 2, fc.get("predicted_revenue", 0));   ws.cell(row=r, column=2).fill = fill
        _money(ws, r, 3, fc.get("predicted_expenses", 0));  ws.cell(row=r, column=3).fill = fill
        _money(ws, r, 4, fc.get("predicted_net_income", 0)); ws.cell(row=r, column=4).fill = fill

        lower = fc.get("revenue_lower")
        upper = fc.get("revenue_upper")
        if lower is not None and upper is not None:
            range_text = f"{lower:,}원 ~ {upper:,}원"
        else:
            range_text = "-"
        _style(ws, r, 5, range_text, fill=fill, font=BODY_FONT, align=CENTER)
        r += 1

    # 요약
    r += 1
    _style(ws, r, 1, "월평균 매출",  font=SUBTOTAL_FONT, align=LEFT)
    _money(ws, r, 2, summary.get("avg_monthly_revenue", 0), is_subtotal=True)
    r += 1
    _style(ws, r, 1, "수익 트렌드", font=SUBTOTAL_FONT, align=LEFT)
    _style(ws, r, 2, summary.get("revenue_trend", ""), font=SUBTOTAL_FONT, align=CENTER)
    r += 1
    _style(ws, r, 1, "예측 모델",   font=SUBTOTAL_FONT, align=LEFT)
    _style(ws, r, 2, model_label,   font=SUBTOTAL_FONT, align=CENTER)


# ═══════════════════════════════════════════════════════════════════
# ⑥ 재고 예측
# ═══════════════════════════════════════════════════════════════════
def _write_inventory_forecast(ws, forecasts: list):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 10

    _title_row(ws, "재고 소진 예측", 10)
    _header_row(ws, [
        "재료명", "카테고리", "현재 재고", "유효 재고",
        "최소 재고", "누적 손실", "일평균 소비", "소진까지(일)",
        "예상 소진일", "상태"
    ], 2)

    STATUS_COLOR = {
        "소진": "E74C3C", "부족": "F39C12", "임박": "F1C40F",
        "주의": "85C1E9", "정상": "82E0AA", "데이터없음": "D5D8DC",
    }

    for i, fc in enumerate(forecasts, 3):
        alt = i % 2 == 0
        base_fill = ALT_FILL if alt else PatternFill()
        status = fc.get("status", "정상")
        stat_fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status, "FFFFFF"))

        vals = [
            fc.get("name", ""),
            fc.get("category", ""),
            fc.get("current_stock", 0),
            fc.get("effective_stock", fc.get("current_stock", 0)),
            fc.get("min_stock", 0),
            fc.get("net_loss", 0),
            fc.get("avg_daily_consumption", 0),
            fc.get("days_until_empty"),
            fc.get("predicted_depletion_date"),
            status,
        ]
        # 상태 컬럼은 J(10번째)
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill   = stat_fill if c == 10 else base_fill
            cell.font   = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c in (2, 8, 9, 10) else (RIGHT if c in (3, 4, 5, 6, 7) else LEFT)
            if c in (3, 4, 5, 6, 7):
                cell.number_format = '#,##0.000'


# ═══════════════════════════════════════════════════════════════════
# 메인 빌더
# ═══════════════════════════════════════════════════════════════════
def build_excel(
    journal:            dict,
    income_statement:   dict,
    balance_sheet:      dict,
    capital_changes:    dict,
    revenue_forecast:   dict,
    inventory_forecast: list,
    filename:           str = None,
) -> str:

    from datetime import datetime
    if not filename:
        filename = f"재무분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()

    # Sheet 1: 손익계산서
    ws1 = wb.active
    ws1.title = "손익계산서"
    _write_income_sheet(ws1, income_statement)

    # Sheet 2: 재무상태표
    ws2 = wb.create_sheet("재무상태표")
    _write_balance_sheet(ws2, balance_sheet)

    # Sheet 3: 자본변동표
    ws3 = wb.create_sheet("자본변동표")
    _write_capital_changes(ws3, capital_changes)

    # Sheet 4: 분개장
    ws4 = wb.create_sheet("분개장")
    _write_journal(ws4, journal)

    # Sheet 5: 수익 예측
    ws5 = wb.create_sheet("수익예측")
    _write_revenue_forecast(ws5, revenue_forecast if isinstance(revenue_forecast, dict) else {})

    # Sheet 6: 재고 예측
    ws6 = wb.create_sheet("재고예측")
    _write_inventory_forecast(ws6, inventory_forecast if isinstance(inventory_forecast, list) else [])

    output_path = EXCEL_OUTPUT_DIR / filename
    wb.save(str(output_path))
    print(f"[excel] 저장 완료: {output_path}")
    return str(output_path)
